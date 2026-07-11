import csv
from io import BytesIO

from django.http import FileResponse, HttpResponse
from django.utils.http import content_disposition_header
from openpyxl import Workbook
from rest_framework.decorators import api_view

from core.bucket.bucketLocal import BucketLocal
from excel.constants import CSV_CONTENT_TYPE, EXCEL_CONTENT_TYPE, EXPORT_COLUMNS
from excel.models import Example, LightweightExample
from excel.serializers import (
    ExampleCSVExportSerializer,
    ExampleExportSerializer,
    LightweightExampleExportSerializer,
)


EXPORT_CHUNK_SIZE = 5_000


@api_view(['GET'])
def export_examples_v1(request):
    serializer = ExampleExportSerializer(data=request.query_params)
    serializer.is_valid(raise_exception=True)

    filename = serializer.validated_data.get('filename')
    filters = {
        column: serializer.validated_data[column]
        for column in EXPORT_COLUMNS
    }
    queryset = Example.objects.filter(**filters).order_by('id')

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = 'Examples'
    worksheet.append(EXPORT_COLUMNS)

    rows = queryset.values_list(*EXPORT_COLUMNS)
    for row in rows:
        worksheet.append(row)

    output = BytesIO()
    workbook.save(output)

    response = HttpResponse(
        output.getvalue(),
        content_type=EXCEL_CONTENT_TYPE,
    )
    response['Content-Disposition'] = content_disposition_header(
        as_attachment=True,
        filename=filename,
    )
    return response


@api_view(['GET'])
def export_examples_v2(request):
    serializer = ExampleCSVExportSerializer(data=request.query_params)
    serializer.is_valid(raise_exception=True)

    filename = serializer.validated_data.get('filename')
    filters = {
        column: serializer.validated_data[column]
        for column in EXPORT_COLUMNS
    }
    queryset = Example.objects.filter(**filters).order_by('id')

    bucket = BucketLocal()
    stored_file = bucket.create_file(filename)

    try:
        with stored_file.path.open(
            'w',
            encoding='utf-8-sig',
            newline='',
        ) as output:
            writer = csv.writer(output)
            writer.writerow(EXPORT_COLUMNS)
            rows = queryset.values_list(*EXPORT_COLUMNS).iterator(
                chunk_size=EXPORT_CHUNK_SIZE,
            )
            writer.writerows(rows)
    except Exception:
        bucket.delete(stored_file.key)
        raise

    return FileResponse(
        bucket.open(stored_file.key),
        as_attachment=True,
        filename=filename,
        content_type=CSV_CONTENT_TYPE,
    )


@api_view(['GET'])
def list_lightweight_examples_v1(request):
    serializer = LightweightExampleExportSerializer(data=request.query_params)
    serializer.is_valid(raise_exception=True)
    filename = serializer.validated_data['filename']

    workbook = Workbook(write_only=True)
    worksheet = workbook.create_sheet(title='Lightweight Examples')
    worksheet.append(('name', 'value'))

    rows = LightweightExample.objects.order_by('id').values_list('name', 'value')
    for row in rows:
        worksheet.append(row)

    output = BytesIO()
    workbook.save(output)

    response = HttpResponse(output.getvalue(), content_type=EXCEL_CONTENT_TYPE)
    response['Content-Disposition'] = content_disposition_header(
        as_attachment=True,
        filename=filename,
    )
    return response
