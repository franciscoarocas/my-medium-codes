import csv
from io import BytesIO
from io import StringIO

from django.test import TestCase
from django.urls import reverse
from openpyxl import load_workbook

from excel.constants import CSV_CONTENT_TYPE, EXCEL_CONTENT_TYPE
from excel.models import Example, LightweightExample


class ExampleExportViewTests(TestCase):
    def test_exports_filtered_examples_as_xlsx(self):
        Example.objects.create(col_a='A1', col_b='B1', col_c='C1', col_d='D1')
        Example.objects.create(col_a='A2', col_b='B2', col_c='C2', col_d='D2')

        response = self.client.get(reverse('export_examples_v1'), {
            'col_a': 'A1',
            'col_b': 'B1',
            'col_c': 'C1',
            'col_d': 'D1',
        })

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response['Content-Type'],
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        self.assertEqual(
            response['Content-Disposition'],
            'attachment; filename="example.xlsx"',
        )

        workbook = load_workbook(BytesIO(response.content))
        worksheet = workbook['Examples']
        rows = list(worksheet.values)

        self.assertEqual(rows[0], ('col_a', 'col_b', 'col_c', 'col_d'))
        self.assertEqual(rows[1], ('A1', 'B1', 'C1', 'D1'))
        self.assertEqual(len(rows), 2)

    def test_uses_filename_query_param_for_download(self):
        Example.objects.create(col_a='A1', col_b='B1', col_c='C1', col_d='D1')

        response = self.client.get(reverse('export_examples_v1'), {
            'col_a': 'A1',
            'col_b': 'B1',
            'col_c': 'C1',
            'col_d': 'D1',
            'filename': 'filtered examples',
        })

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response['Content-Disposition'],
            'attachment; filename="filtered_examples.xlsx"',
        )

    def test_rejects_filename_over_limit(self):
        response = self.client.get(reverse('export_examples_v1'), {
            'col_a': 'A1',
            'col_b': 'B1',
            'col_c': 'C1',
            'col_d': 'D1',
            'filename': 'a' * 101,
        })

        self.assertEqual(response.status_code, 400)
        self.assertIn('filename', response.json())


class ExampleExportV2ViewTests(TestCase):
    def test_streams_filtered_examples_from_a_bucket_csv_file(self):
        Example.objects.create(col_a='A1', col_b='B1', col_c='C1', col_d='D1')
        Example.objects.create(col_a='A2', col_b='B2', col_c='C2', col_d='D2')

        response = self.client.get(reverse('export_examples_v2'), {
            'col_a': 'A1',
            'col_b': 'B1',
            'col_c': 'C1',
            'col_d': 'D1',
            'filename': 'streamed examples',
        })
        content = b''.join(response.streaming_content)

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response['Content-Disposition'],
            'attachment; filename="streamed_examples.csv"',
        )
        self.assertEqual(response['Content-Type'], CSV_CONTENT_TYPE)
        rows = list(csv.reader(StringIO(content.decode('utf-8-sig'))))
        self.assertEqual(rows, [
            ['col_a', 'col_b', 'col_c', 'col_d'],
            ['A1', 'B1', 'C1', 'D1'],
        ])


class LightweightExampleExportViewTests(TestCase):
    def test_exports_lightweight_examples_as_xlsx(self):
        LightweightExample.objects.create(name='item-01', value=40)
        LightweightExample.objects.create(name='item-02', value=80)

        response = self.client.get(reverse('list_lightweight_examples_v1'))

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response['Content-Type'], EXCEL_CONTENT_TYPE)
        self.assertEqual(
            response['Content-Disposition'],
            'attachment; filename="lightweight_examples.xlsx"',
        )

        workbook = load_workbook(BytesIO(response.content))
        rows = list(workbook['Lightweight Examples'].values)
        self.assertEqual(rows, [
            ('name', 'value'),
            ('item-01', 40),
            ('item-02', 80),
        ])

    def test_uses_filename_query_param_for_download(self):
        response = self.client.get(
            reverse('list_lightweight_examples_v1'),
            {'filename': 'small export'},
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response['Content-Disposition'],
            'attachment; filename="small_export.xlsx"',
        )
